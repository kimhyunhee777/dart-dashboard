# -*- coding: utf-8 -*-
"""
DART(전자공시시스템) Open API 기반 재무분석기
- 회사명으로 검색
- 최대 3개 회사 x 3개년 재무제표 비교
- 결과를 엑셀 파일로 저장
"""

import io
import json
import sys
import zipfile
import xml.etree.ElementTree as ET
from datetime import datetime
from pathlib import Path

import os

import requests
import pandas as pd
from dotenv import load_dotenv
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Windows 콘솔(cp949 등)에서 한글이 깨지는 것을 방지
try:
    sys.stdin.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass
for _stream in (sys.stdout, sys.stderr):
    try:
        _stream.reconfigure(encoding="utf-8")
    except Exception:
        pass

load_dotenv()
openapi = os.environ.get("DART_API_KEY")
if not openapi:
    sys.exit(
        "[오류] DART_API_KEY가 설정되지 않았습니다.\n"
        "  .env 파일을 만들고 다음처럼 입력하세요:\n"
        "  DART_API_KEY=발급받은_API_키"
    )
BASE_URL = "https://opendart.fss.or.kr/api"
SCRIPT_DIR = Path(__file__).resolve().parent
CORP_CODE_CACHE = SCRIPT_DIR / "corpCode_cache.json"

# 재무제표에서 뽑아올 핵심 계정과목 (계정명 매칭용)
TARGET_ACCOUNTS = {
    "매출액": ["매출액", "수익(매출액)", "영업수익"],
    "영업이익": ["영업이익", "영업이익(손실)"],
    "당기순이익": ["당기순이익", "당기순이익(손실)", "분기순이익", "분기순이익(손실)"],
    "자산총계": ["자산총계"],
    "부채총계": ["부채총계"],
    "자본총계": ["자본총계"],
}

REPORT_CODE_ANNUAL = "11011"  # 사업보고서(연간)


def _get(url, params, binary=False):
    resp = requests.get(url, params=params, timeout=30)
    resp.raise_for_status()
    return resp.content if binary else resp.json()


def load_corp_codes(force_refresh=False):
    """DART 전체 고유번호(corp_code) 목록을 다운로드/캐시하여 반환한다."""
    if CORP_CODE_CACHE.exists() and not force_refresh:
        with open(CORP_CODE_CACHE, "r", encoding="utf-8") as f:
            return json.load(f)

    print("[정보] 회사 고유코드 목록을 DART에서 내려받는 중... (최초 1회, 잠시 걸릴 수 있어요)")
    content = _get(f"{BASE_URL}/corpCode.xml", {"crtfc_key": openapi}, binary=True)

    try:
        with zipfile.ZipFile(io.BytesIO(content)) as zf:
            xml_bytes = zf.read("CORPCODE.xml")
    except zipfile.BadZipFile:
        # API 키 오류 등으로 zip이 아니라 에러 JSON이 온 경우
        try:
            err = json.loads(content.decode("utf-8"))
        except Exception:
            err = {"message": content[:200]}
        raise RuntimeError(f"고유코드 다운로드 실패: {err}")

    root = ET.fromstring(xml_bytes)
    corp_list = []
    for item in root.findall("list"):
        corp_list.append({
            "corp_code": (item.findtext("corp_code") or "").strip(),
            "corp_name": (item.findtext("corp_name") or "").strip(),
            "stock_code": (item.findtext("stock_code") or "").strip(),
            "modify_date": (item.findtext("modify_date") or "").strip(),
        })

    with open(CORP_CODE_CACHE, "w", encoding="utf-8") as f:
        json.dump(corp_list, f, ensure_ascii=False)

    print(f"[정보] 회사 {len(corp_list)}건 캐시 완료 -> {CORP_CODE_CACHE.name}")
    return corp_list


def search_company(name, corp_list):
    """이름에 검색어가 포함된 회사 목록을 반환. 상장사(stock_code 있음) 우선 정렬."""
    name = name.strip()
    matches = [c for c in corp_list if name in c["corp_name"]]
    matches.sort(key=lambda c: (c["stock_code"] == "", c["corp_name"]))
    return matches


def pick_company(name, corp_list):
    matches = search_company(name, corp_list)
    if not matches:
        print(f"  -> '{name}' 검색 결과 없음")
        return None

    listed = [m for m in matches if m["stock_code"]]
    candidates = listed if listed else matches

    if len(candidates) == 1:
        chosen = candidates[0]
        print(f"  -> '{name}' 검색 결과: {chosen['corp_name']} (종목코드: {chosen['stock_code'] or '비상장'})")
        return chosen

    print(f"  -> '{name}' 검색 결과가 여러 건입니다. 번호를 선택하세요:")
    for i, c in enumerate(candidates[:15], start=1):
        print(f"     [{i}] {c['corp_name']}  (종목코드: {c['stock_code'] or '비상장'})")
    while True:
        sel = input("     선택 번호 입력 (기본 1): ").strip()
        if sel == "":
            return candidates[0]
        if sel.isdigit() and 1 <= int(sel) <= len(candidates[:15]):
            return candidates[int(sel) - 1]
        print("     잘못된 입력입니다. 다시 입력하세요.")


def fetch_financial_statement(corp_code, year, reprt_code=REPORT_CODE_ANNUAL):
    """단일회사 전체 재무제표 조회. 연결(CFS) 우선, 없으면 개별(OFS) 재시도."""
    for fs_div in ("CFS", "OFS"):
        params = {
            "crtfc_key": openapi,
            "corp_code": corp_code,
            "bsns_year": str(year),
            "reprt_code": reprt_code,
            "fs_div": fs_div,
        }
        data = _get(f"{BASE_URL}/fnlttSinglAcntAll.json", params)
        if data.get("status") == "000":
            return data.get("list", []), fs_div
    return [], None


def extract_metrics(rows):
    """계정과목 리스트에서 핵심 지표만 추출 (재무상태표/손익계산서 통합)."""
    result = {k: None for k in TARGET_ACCOUNTS}
    for row in rows:
        account_nm = (row.get("account_nm") or "").strip()
        amount_str = (row.get("thstrm_amount") or "").replace(",", "").strip()
        if not amount_str:
            continue
        try:
            amount = float(amount_str)
        except ValueError:
            continue
        for metric, aliases in TARGET_ACCOUNTS.items():
            if result[metric] is not None:
                continue
            if account_nm in aliases:
                result[metric] = amount
    return result


def collect_data(company_names, years):
    corp_list = load_corp_codes()

    resolved = []
    print("\n[회사 검색]")
    for name in company_names:
        chosen = pick_company(name, corp_list)
        if chosen:
            resolved.append(chosen)
    if not resolved:
        print("검색된 회사가 없어 종료합니다.")
        sys.exit(1)

    records = []
    detail_rows = []
    print("\n[재무제표 조회]")
    for corp in resolved:
        for year in years:
            print(f"  - {corp['corp_name']} / {year}년 조회 중...")
            try:
                rows, fs_div = fetch_financial_statement(corp["corp_code"], year)
            except requests.RequestException as e:
                print(f"    [오류] 네트워크 문제: {e}")
                continue

            if not rows:
                print(f"    [알림] {year}년 데이터 없음 (미공시 또는 사업보고서 미제출)")
                continue

            metrics = extract_metrics(rows)
            record = {"회사명": corp["corp_name"], "종목코드": corp["stock_code"] or "비상장",
                      "연도": year, "재무제표구분": "연결" if fs_div == "CFS" else "개별"}
            record.update(metrics)

            # 비율 계산
            revenue = metrics.get("매출액")
            op_income = metrics.get("영업이익")
            net_income = metrics.get("당기순이익")
            assets = metrics.get("자산총계")
            liabilities = metrics.get("부채총계")
            equity = metrics.get("자본총계")

            record["영업이익률(%)"] = round(op_income / revenue * 100, 2) if revenue and op_income is not None else None
            record["순이익률(%)"] = round(net_income / revenue * 100, 2) if revenue and net_income is not None else None
            record["부채비율(%)"] = round(liabilities / equity * 100, 2) if equity and liabilities is not None else None
            record["ROE(%)"] = round(net_income / equity * 100, 2) if equity and net_income is not None else None
            record["ROA(%)"] = round(net_income / assets * 100, 2) if assets and net_income is not None else None

            records.append(record)

            for row in rows:
                detail_rows.append({
                    "회사명": corp["corp_name"], "연도": year, "재무제표구분": "연결" if fs_div == "CFS" else "개별",
                    "재무제표명": row.get("sj_nm"), "계정명": row.get("account_nm"),
                    "당기금액": row.get("thstrm_amount"), "전기금액": row.get("frmtrm_amount"),
                    "전전기금액": row.get("bfefrmtrm_amount"),
                })

    return records, detail_rows


def format_money(v):
    if v is None:
        return None
    return v  # 숫자 그대로 저장, 엑셀 서식으로 표시


def build_excel(records, detail_rows, output_path):
    if not records:
        print("저장할 데이터가 없습니다.")
        return

    summary_df = pd.DataFrame(records)
    money_cols = ["매출액", "영업이익", "당기순이익", "자산총계", "부채총계", "자본총계"]
    summary_df = summary_df.sort_values(["회사명", "연도"]).reset_index(drop=True)

    detail_df = pd.DataFrame(detail_rows)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="요약비교", index=False)
        if not detail_df.empty:
            detail_df.to_excel(writer, sheet_name="원본데이터", index=False)

        # 회사별 피벗(연도별 비교가 한눈에 보이도록)
        for metric in money_cols + ["영업이익률(%)", "순이익률(%)", "부채비율(%)", "ROE(%)", "ROA(%)"]:
            if metric not in summary_df.columns:
                continue
            pivot = summary_df.pivot_table(index="회사명", columns="연도", values=metric, aggfunc="first")
            sheet_name = f"{metric[:20]}"
            # 시트명 특수문자 제거
            for ch in "[]:*?/\\":
                sheet_name = sheet_name.replace(ch, "")
            pivot.to_excel(writer, sheet_name=sheet_name)

    # ---- 서식 다듬기 ----
    from openpyxl import load_workbook
    wb = load_workbook(output_path)

    header_fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for ws in wb.worksheets:
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.freeze_panes = "A2"

        # 열 너비 자동 조정
        for col_idx, col_cells in enumerate(ws.iter_cols(min_row=1, max_row=ws.max_row), start=1):
            max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 40)

        # 금액 열 천단위 콤마 서식 (요약비교 시트)
        if ws.title == "요약비교":
            header_row = [c.value for c in ws[1]]
            for col_name in money_cols:
                if col_name in header_row:
                    col_idx = header_row.index(col_name) + 1
                    for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                        for cell in row:
                            cell.number_format = "#,##0"

    wb.save(output_path)
    print(f"\n[완료] 엑셀 파일 저장: {output_path}")


def main():
    print("=" * 50)
    print(" DART 재무분석기 - 회사명 검색 및 3개년 재무제표 비교")
    print("=" * 50)

    companies = []
    print("\n비교할 회사명을 최대 3개까지 입력하세요 (예: 삼성전자). 그만 입력하려면 Enter.")
    for i in range(1, 4):
        name = input(f"  회사 {i} 이름: ").strip()
        if not name:
            break
        companies.append(name)

    if not companies:
        print("입력된 회사가 없어 종료합니다.")
        return

    this_year = datetime.now().year
    default_end_year = this_year - 1  # 사업보고서는 통상 익년 3월 제출
    year_input = input(f"\n조회 종료연도 입력 (기본 {default_end_year}, 최근 3개년을 조회합니다): ").strip()
    end_year = int(year_input) if year_input.isdigit() else default_end_year
    years = [end_year - 2, end_year - 1, end_year]

    records, detail_rows = collect_data(companies, years)

    output_path = SCRIPT_DIR / f"재무분석_비교결과_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    build_excel(records, detail_rows, output_path)


if __name__ == "__main__":
    main()
